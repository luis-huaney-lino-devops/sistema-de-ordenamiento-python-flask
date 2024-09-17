from flask import Flask, render_template, request, redirect, url_for, jsonify, send_from_directory, Response
import os
import random
import string
import time
import threading
from faker import Faker
from openpyxl import load_workbook, Workbook
from werkzeug.utils import secure_filename
app = Flask(__name__)
# Configuración de carpetas
# Configuración de carpetas
app.config['carpeta_archivos_subidos'] = 'archivos_subidos'
app.config['carpeta_burble'] = 'archivos_burble_sort'
app.config['carpeta_insertion'] = 'archivos_insertion_sort'
app.config['carpeta_selection'] = 'archivos_selection_sort'
app.config['carpeta_merge'] = 'archivos_merge_sort'
app.config['carpeta_heap'] = 'archivos_heap_sort'
app.config['carpeta_counting'] = 'archivos_counting_sort'
app.config['carpeta_radix'] = 'archivos_radix_sort'
app.config['carpeta_quick'] = 'archivos_quick_sort'


# Creación de carpetas si no existen
for carpeta in [
    'carpeta_archivos_subidos', 'carpeta_burble', 'carpeta_insertion',
    'carpeta_selection', 'carpeta_merge', 'carpeta_heap', 'carpeta_counting',
    'carpeta_radix', 'carpeta_quick'
]:
    if not os.path.exists(app.config[carpeta]):
        os.makedirs(app.config[carpeta])

# Variables de progreso
progress_burble_sort_arr = {'progress_burble_sort': 0}
progress_insertion_sort_arr = {'progress_insertion_sort': 0}
progress_selection_sort_arr = {'progress_selection_sort': 0}
progress_merge_sort_arr = {'progress_merge_sort': 0}
progress_heap_sort_arr = {'progress_heap_sort': 0}
progress_counting_sort_arr = {'progress_counting_sort': 0}
progress_radix_sort_arr = {'progress_radix_sort': 0}
progress_quick_sort_arr = {'progress_quick_sort': 0}

def generate_unique_filename(base_path, base_filename, extension):
    counter = 1
    filename = f"{base_filename}.{extension}"
    while os.path.exists(os.path.join(base_path, filename)):
        filename = f"{base_filename}_{counter}.{extension}"
        counter += 1
    return filename

def bubble_sort(numbers):
    n = len(numbers)
    for i in range(n):
        for j in range(0, n - i - 1):
            if numbers[j] > numbers[j + 1]:
                numbers[j], numbers[j + 1] = numbers[j + 1], numbers[j]
            progress_burble_sort_arr['progress_burble_sort'] = int(((i * n + j + 1) / (n * n)) * 100)
            time.sleep(0.01)  # Simulate time-consuming task
    progress_burble_sort_arr['progress_burble_sort'] = 100
    return numbers

def insertion_sort(numbers):
    n = len(numbers)
    for i in range(1, n):
        key = numbers[i]
        j = i - 1
        while j >= 0 and key < numbers[j]:
            numbers[j + 1] = numbers[j]
            j -= 1
        numbers[j + 1] = key
        progress_insertion_sort_arr['progress_insertion_sort'] = int(((i + 1) / n) * 100)
        time.sleep(0.01)  # Simulate time-consuming task
    progress_insertion_sort_arr['progress_insertion_sort'] = 100
    return numbers

def selection_sort(numbers):
    n = len(numbers)
    for i in range(n):
        min_idx = i
        for j in range(i + 1, n):
            if numbers[min_idx] > numbers[j]:
                min_idx = j
        numbers[i], numbers[min_idx] = numbers[min_idx], numbers[i]
        progress_selection_sort_arr['progress_selection_sort'] = int(((i + 1) / n) * 100)
        time.sleep(0.01)  # Simulate time-consuming task
    progress_selection_sort_arr['progress_selection_sort'] = 100
    return numbers

def merge_sort(numbers):
    def merge(left, right):
        result = []
        left_idx, right_idx = 0, 0
        while left_idx < len(left) and right_idx < len(right):
            if left[left_idx] < right[right_idx]:
                result.append(left[left_idx])
                left_idx += 1
            else:
                result.append(right[right_idx])
                right_idx += 1
        result.extend(left[left_idx:])
        result.extend(right[right_idx:])
        return result
    
    if len(numbers) <= 1:
        return numbers
    mid = len(numbers) // 2
    left = merge_sort(numbers[:mid])
    right = merge_sort(numbers[mid:])
    sorted_numbers = merge(left, right)
    progress_merge_sort_arr['progress_merge_sort'] = int((len(sorted_numbers) / len(numbers)) * 100)
    time.sleep(0.01)  # Simulate time-consuming task
    return sorted_numbers

def heap_sort(numbers):
    def heapify(arr, n, i):
        largest = i
        l = 2 * i + 1
        r = 2 * i + 2
        if l < n and arr[i] < arr[l]:
            largest = l
        if r < n and arr[largest] < arr[r]:
            largest = r
        if largest != i:
            arr[i], arr[largest] = arr[largest], arr[i]
            heapify(arr, n, largest)
    
    n = len(numbers)
    for i in range(n // 2 - 1, -1, -1):
        heapify(numbers, n, i)
    for i in range(n - 1, 0, -1):
        numbers[i], numbers[0] = numbers[0], numbers[i]
        heapify(numbers, i, 0)
        progress_heap_sort_arr['progress_heap_sort'] = int(((n - i) / n) * 100)
        time.sleep(0.01)  # Simulate time-consuming task
    progress_heap_sort_arr['progress_heap_sort'] = 100
    return numbers

def counting_sort(numbers):
    max_val = max(numbers)
    m = max_val + 1
    count = [0] * m
    for a in numbers:
        count[a] += 1
    i = 0
    for a in range(m):
        for c in range(count[a]):
            numbers[i] = a
            i += 1
        progress_counting_sort_arr['progress_counting_sort'] = int((i / len(numbers)) * 100)
        time.sleep(0.01)  # Simulate time-consuming task
    progress_counting_sort_arr['progress_counting_sort'] = 100
    return numbers

def radix_sort(numbers):
    RADIX = 10
    placement = 1
    max_digit = max(numbers)
    while placement < max_digit:
        buckets = [list() for _ in range(RADIX)]
        for i in numbers:
            tmp = int((i / placement) % RADIX)
            buckets[tmp].append(i)
        a = 0
        for b in range(RADIX):
            buck = buckets[b]
            for i in buck:
                numbers[a] = i
                a += 1
        placement *= RADIX
        progress_radix_sort_arr['progress_radix_sort'] = int((placement / max_digit) * 100)
        time.sleep(0.01)  # Simulate time-consuming task
    progress_radix_sort_arr['progress_radix_sort'] = 100
    return numbers

def quick_sort(numbers):
    def partition(arr, low, high):
        i = (low - 1)
        pivot = arr[high]
        for j in range(low, high):
            if arr[j] < pivot:
                i = i + 1
                arr[i], arr[j] = arr[j], arr[i]
        arr[i + 1], arr[high] = arr[high], arr[i + 1]
        return (i + 1)
    
    def quick_sort_recursive(arr, low, high):
        if low < high:
            pi = partition(arr, low, high)
            quick_sort_recursive(arr, low, pi - 1)
            quick_sort_recursive(arr, pi + 1, high)
            progress_quick_sort_arr['progress_quick_sort'] = int(((high - low + 1) / len(numbers)) * 100)
            time.sleep(0.01)  # Simulate time-consuming task
    
    quick_sort_recursive(numbers, 0, len(numbers) - 1)
    progress_quick_sort_arr['progress_quick_sort'] = 100
    return numbers

@app.route('/')
def index():
    return render_template('inicio.html')

@app.route('/archivos_burble_sort/')
def list_archivos_burble_sort():
    # Lista de archivos en la carpeta archivos_burble_sort
    files = os.listdir(app.config['carpeta_burble'])
    return render_template('lista_archivos.html', files=files)

@app.route('/archivos_burble_sort/<filename>')
def download_file(filename):
    # Descargar el archivo seleccionado
    return send_from_directory(app.config['carpeta_burble'], filename)

@app.route('/burble')
def burble():
    return render_template('burble.html')
  
@app.route('/burble_teori')
def burble_teori():
    return render_template('burble_teori.html')   

@app.route('/burble_archi')
def burble_archi():
    return render_template('burble_archi.html')

@app.route('/upload_burble_archiv', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['carpeta_archivos_subidos'], filename)
        file.save(filepath)

        filename, file_extension = os.path.splitext(file.filename)
        numbers = []

        if file_extension in ['.txt', '.csv']:
            with open(filepath, 'r') as f:
                data = f.read()
            numbers = list(map(int, data.split(',')))

        elif file_extension == '.xlsx':
            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, int):
                        numbers.append(cell)
        else:
            return jsonify({'error': 'Unsupported file type'}), 400

        sorting_thread = threading.Thread(target=bubble_sort, args=(numbers,))
        sorting_thread.start()
        sorting_thread.join()

        sorted_filename = generate_unique_filename(app.config['carpeta_burble'], f"Burble_sort_{filename}", file_extension[1:])
        sorted_filepath = os.path.join(app.config['carpeta_burble'], sorted_filename)

        if file_extension in ['.txt', '.csv']:
            with open(sorted_filepath, 'w') as f:
                f.write(','.join(map(str, numbers)))

        elif file_extension == '.xlsx':
            wb = Workbook()
            ws = wb.active
            for idx, num in enumerate(numbers, start=1):
                ws.cell(row=idx, column=1, value=num)
            wb.save(sorted_filepath)

        return jsonify({
            'success': True,
            'download_url': url_for('download_file', filename=sorted_filename)
        })




@app.route('/progress')
def progress():
    def generate():
        while progress_burble_sort_arr['progress_burble_sort'] < 100:
            yield f"data: {progress_burble_sort_arr['progress_burble_sort']}\n\n"
            time.sleep(1)
        yield f"data: 100\n\n"
    return Response(generate(), mimetype='text/event-stream')

@app.route('/burble_gen_ale')

def burble_gen_ale():
    return render_template('burble_gen_ale.html')

fake = Faker()

@app.route('/upload_burble_gen_ale', methods=['POST'])
def upload_burble_gen_ale():
    tipo_datos = request.form['tipo_datos']
    cantidad = int(request.form['cantidad_date'])

    if tipo_datos == 'numeros_enteros':
        generated_data = [random.randint(0, 100) for _ in range(cantidad)]
    elif tipo_datos == 'numeros_decimales':
        generated_data = [round(random.uniform(0, 100), 2) for _ in range(cantidad)]
    elif tipo_datos == 'nombres':
        generated_data = [fake.first_name() for _ in range(cantidad)]
    elif tipo_datos == 'palabras':
        generated_data = [fake.word() for _ in range(cantidad)]
    else:
        return jsonify({'error': 'Unsupported data type'}), 400

    return jsonify({'success': True, 'generated_data': generated_data})
  
  
@app.route('/ordenacion_burbuja_datos_aleatorios', methods=['POST'])
def ordenacion_burbuja_datos_aleatorios():
    data = request.json['data']
    # Ordenar los datos usando el algoritmo de ordenación burbuja
    sorted_data = bubble_sort(data)

    # Generar un nombre de archivo único
    base_filename = 'datos_aletorios_ordenados_burble'
    sorted_filename = generate_unique_filename(app.config['carpeta_burble'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_burble'], sorted_filename)

    # Guardar los datos ordenados en un archivo
    with open(sorted_filepath, 'w') as f:
        f.write(','.join(map(str, sorted_data)))

    # Devolver la respuesta con los datos ordenados y la URL de descarga
    return jsonify({'success': True, 'sorted_data': sorted_data, 'download_url': url_for('download_file', filename=sorted_filename)})



@app.route('/burble_manual')
def burble_manual():
    return render_template('burble_manual.html')   


@app.route('/burble_manual_form', methods=['POST'])
def burble_manual_form():
    tipo_datos = request.form['tipo_datos']
    datos_a_ordenar = request.form['date_manu']
    separador = request.form['separador']

    if separador == 'comas':
        datos_list = datos_a_ordenar.split(',')
    elif separador == 'espacios':
        datos_list = datos_a_ordenar.split()

    if tipo_datos == 'numeros_enteros':
        datos_list = list(map(int, datos_list))
    elif tipo_datos == 'palabras':
        datos_list = list(map(str, datos_list))

    sorting_thread = threading.Thread(target=bubble_sort, args=(datos_list,))
    sorting_thread.start()
    sorting_thread.join()

    sorted_list = datos_list
    sorted_data = ', '.join(map(str, sorted_list)) if separador == 'comas' else ' '.join(map(str, sorted_list))

    # Generar un nombre de archivo único
    base_filename = 'datos_manualmente_ordenados'
    sorted_filename = generate_unique_filename(app.config['carpeta_burble'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_burble'], sorted_filename)

    with open(sorted_filepath, 'w') as f:
        f.write(sorted_data)

    return jsonify({
        'sorted_data': sorted_data,
        'download_url': url_for('download_file', filename=sorted_filename)
    })
    



@app.route('/archivos_insertion_sort/')
def list_archivos_insertion_sort():
    # Lista de archivos en la carpeta archivos_burble_sort
    files = os.listdir(app.config['carpeta_insertion'])
    return render_template('lista_archivos.html', files=files)

@app.route('/archivos_insertion_sort/<filename>')
def download_file_insert(filename):
    # Descargar el archivo seleccionado
    return send_from_directory(app.config['carpeta_insertion'], filename)

@app.route('/insertion')
def insertion():
    return render_template('insertion.html')  
   
@app.route('/insertion_teori')
def insertion_teori():
    return render_template('insertion_teori.html')   
@app.route('/insertion_archi')
def insertion_archi():
    return render_template('insertion_archi.html')
  
@app.route('/upload_insertion_archiv', methods=['POST'])
def upload_insertion_archiv():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['carpeta_archivos_subidos'], filename)
        file.save(filepath)

        filename, file_extension = os.path.splitext(file.filename)
        numbers = []

        if file_extension in ['.txt', '.csv']:
            with open(filepath, 'r') as f:
                data = f.read()
            numbers = list(map(int, data.split(',')))

        elif file_extension == '.xlsx':
            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, int):
                        numbers.append(cell)
        else:
            return jsonify({'error': 'Unsupported file type'}), 400

        sorting_thread = threading.Thread(target=insertion_sort, args=(numbers,))
        sorting_thread.start()
        sorting_thread.join()

        sorted_filename = generate_unique_filename(app.config['carpeta_burble'], f"Insert_sort_{filename}", file_extension[1:])
        sorted_filepath = os.path.join(app.config['carpeta_burble'], sorted_filename)

        if file_extension in ['.txt', '.csv']:
            with open(sorted_filepath, 'w') as f:
                f.write(','.join(map(str, numbers)))

        elif file_extension == '.xlsx':
            wb = Workbook()
            ws = wb.active
            for idx, num in enumerate(numbers, start=1):
                ws.cell(row=idx, column=1, value=num)
            wb.save(sorted_filepath)

        return jsonify({
            'success': True,
            'download_url': url_for('download_file', filename=sorted_filename)
        })

@app.route('/progress_insertion')
def progress_insertion():
    def generate():
        while progress_insertion_sort_arr['progress_insertion_sort'] < 100:
            yield f"data: {progress_insertion_sort_arr['progress_insertion_sort']}\n\n"
            time.sleep(1)
        yield f"data: 100\n\n"
    return Response(generate(), mimetype='text/event-stream')

@app.route('/insertion_gen_ale')
def insertion_gen_ale():
    return render_template('insertion_gen_ale.html')

@app.route('/upload_insert_gen_ale', methods=['POST'])
def upload_insert_gen_ale():
    tipo_datos = request.form['tipo_datos']
    cantidad = int(request.form['cantidad_date'])

    if tipo_datos == 'numeros_enteros':
        generated_data = [random.randint(0, 100) for _ in range(cantidad)]
    elif tipo_datos == 'numeros_decimales':
        generated_data = [round(random.uniform(0, 100), 2) for _ in range(cantidad)]
    elif tipo_datos == 'nombres':
        generated_data = [fake.first_name() for _ in range(cantidad)]
    elif tipo_datos == 'palabras':
        generated_data = [fake.word() for _ in range(cantidad)]
    else:
        return jsonify({'error': 'Unsupported data type'}), 400

    return jsonify({'success': True, 'generated_data': generated_data})

@app.route('/ordenacion_insert_datos_aleatorios', methods=['POST'])
def ordenacion_insert_datos_aleatorios():
    data = request.json['data']
    # Ordenar los datos usando el algoritmo de ordenación burbuja
    sorted_data = insertion_sort(data)

    # Generar un nombre de archivo único
    base_filename = 'datos_aletorios_ordenados_insert'
    sorted_filename = generate_unique_filename(app.config['carpeta_insertion'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_insertion'], sorted_filename)

    # Guardar los datos ordenados en un archivo
    with open(sorted_filepath, 'w') as f:
        f.write(','.join(map(str, sorted_data)))

    # Devolver la respuesta con los datos ordenados y la URL de descarga
    return jsonify({'success': True, 'sorted_data': sorted_data, 'download_url': url_for('download_file', filename=sorted_filename)})


@app.route('/insertion_manual')
def insertion_manual():
    return render_template('insertion_manual.html')   

@app.route('/insertion_manual_form', methods=['POST'])
def insertion_manual_form():
    tipo_datos = request.form['tipo_datos']
    datos_a_ordenar = request.form['date_manu']
    separador = request.form['separador']

    if separador == 'comas':
        datos_list = datos_a_ordenar.split(',')
    elif separador == 'espacios':
        datos_list = datos_a_ordenar.split()

    if tipo_datos == 'numeros_enteros':
        datos_list = list(map(int, datos_list))
    elif tipo_datos == 'palabras':
        datos_list = list(map(str, datos_list))

    sorting_thread = threading.Thread(target=insertion_sort, args=(datos_list,))
    sorting_thread.start()
    sorting_thread.join()

    sorted_list = datos_list
    sorted_data = ', '.join(map(str, sorted_list)) if separador == 'comas' else ' '.join(map(str, sorted_list))
    # Generar un nombre de archivo único
    base_filename = 'datos_manualmente_ordenados'
    sorted_filename = generate_unique_filename(app.config['carpeta_burble'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_burble'], sorted_filename)

    with open(sorted_filepath, 'w') as f:
        f.write(sorted_data)

    return jsonify({
        'sorted_data': sorted_data,
        'download_url': url_for('download_file', filename=sorted_filename)
    })


@app.route('/archivos_selection_sort/')
def list_archivos_selection_sort():
    # Lista de archivos en la carpeta archivos_burble_sort
    files = os.listdir(app.config['carpeta_selection'])
    return render_template('lista_archivos_select.html', files=files)

@app.route('/archivos_selection_sort/<filename>')
def download_file_selection(filename):
    # Descargar el archivo seleccionado
    return send_from_directory(app.config['carpeta_selection'], filename)

@app.route('/selection')
def selection():
    return render_template('selection.html')
  
@app.route('/selection_teori')
def selection_teori():
    return render_template('selection_teori.html')

@app.route('/selection_archi')
def selection_archi():
    return render_template('selection_archi.html')

@app.route('/upload_selection_archiv', methods=['POST'])
def upload_selection_archiv():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['carpeta_archivos_subidos'], filename)
        file.save(filepath)

        filename, file_extension = os.path.splitext(file.filename)
        numbers = []

        if file_extension in ['.txt', '.csv']:
            with open(filepath, 'r') as f:
                data = f.read()
            numbers = list(map(int, data.split(',')))

        elif file_extension == '.xlsx':
            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, int):
                        numbers.append(cell)
        else:
            return jsonify({'error': 'Unsupported file type'}), 400

        sorting_thread = threading.Thread(target=insertion_sort, args=(numbers,))
        sorting_thread.start()
        sorting_thread.join()

        sorted_filename = generate_unique_filename(app.config['carpeta_selection'], f"Selection_sort_{filename}", file_extension[1:])
        sorted_filepath = os.path.join(app.config['carpeta_selection'], sorted_filename)

        if file_extension in ['.txt', '.csv']:
            with open(sorted_filepath, 'w') as f:
                f.write(','.join(map(str, numbers)))

        elif file_extension == '.xlsx':
            wb = Workbook()
            ws = wb.active
            for idx, num in enumerate(numbers, start=1):
                ws.cell(row=idx, column=1, value=num)
            wb.save(sorted_filepath)

        return jsonify({
            'success': True,
            'download_url': url_for('download_file', filename=sorted_filename)
        })

@app.route('/progress_selection')
def progress_selection():
    def generate():
        while progress_selection_sort_arr['progress_selection_sort'] < 100:
            yield f"data: {progress_selection_sort_arr['progress_selection_sort']}\n\n"
            time.sleep(1)
        yield f"data: 100\n\n"
    return Response(generate(), mimetype='text/event-stream')


@app.route('/selection_gen_ale')
def selection_gen_ale():
    return render_template('selection_gen_ale.html')

@app.route('/upload_selection_gen_ale', methods=['POST'])
def upload_selection_gen_ale():
    tipo_datos = request.form['tipo_datos']
    cantidad = int(request.form['cantidad_date'])

    if tipo_datos == 'numeros_enteros':
        generated_data = [random.randint(0, 100) for _ in range(cantidad)]
    elif tipo_datos == 'numeros_decimales':
        generated_data = [round(random.uniform(0, 100), 2) for _ in range(cantidad)]
    elif tipo_datos == 'nombres':
        generated_data = [fake.first_name() for _ in range(cantidad)]
    elif tipo_datos == 'palabras':
        generated_data = [fake.word() for _ in range(cantidad)]
    else:
        return jsonify({'error': 'Unsupported data type'}), 400

    return jsonify({'success': True, 'generated_data': generated_data})

@app.route('/ordenacion_selection_datos_aleatorios', methods=['POST'])
def ordenacion_selection_datos_aleatorios():
    data = request.json['data']
    # Ordenar los datos usando el algoritmo de ordenación burbuja
    sorted_data = selection_sort(data)

    # Generar un nombre de archivo único
    base_filename = 'datos_aletorios_ordenados_selection'
    sorted_filename = generate_unique_filename(app.config['carpeta_selection'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_selection'], sorted_filename)

    # Guardar los datos ordenados en un archivo
    with open(sorted_filepath, 'w') as f:
        f.write(','.join(map(str, sorted_data)))

    # Devolver la respuesta con los datos ordenados y la URL de descarga
    return jsonify({'success': True, 'sorted_data': sorted_data, 'download_url': url_for('download_file', filename=sorted_filename)})

@app.route('/selection_manual')
def selection_manual():
    return render_template('selection_manual.html')   

@app.route('/selection_manual_form', methods=['POST'])
def selection_manual_form():
    tipo_datos = request.form['tipo_datos']
    datos_a_ordenar = request.form['date_manu']
    separador = request.form['separador']

    if separador == 'comas':
        datos_list = datos_a_ordenar.split(',')
    elif separador == 'espacios':
        datos_list = datos_a_ordenar.split()

    if tipo_datos == 'numeros_enteros':
        datos_list = list(map(int, datos_list))
    elif tipo_datos == 'palabras':
        datos_list = list(map(str, datos_list))

    sorting_thread = threading.Thread(target=selection_sort, args=(datos_list,))
    sorting_thread.start()
    sorting_thread.join()

    sorted_list = datos_list
    sorted_data = ', '.join(map(str, sorted_list)) if separador == 'comas' else ' '.join(map(str, sorted_list))
    # Generar un nombre de archivo único
    base_filename = 'datos_manualmente_ordenados_selection'
    sorted_filename = generate_unique_filename(app.config['carpeta_selection'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_selection'], sorted_filename)

    with open(sorted_filepath, 'w') as f:
        f.write(sorted_data)

    return jsonify({
        'sorted_data': sorted_data,
        'download_url': url_for('download_file', filename=sorted_filename)
    })


@app.route('/busqueda_lineal_binaria')
def linear_search_view():
    return render_template('busqueda_lineal_binaria.html')
  
@app.route('/busqueda_lineal')
def busqueda_lineal():
    return render_template('busqueda_lineal.html')
  
@app.route('/busqueda_lineal_teori')
def busqueda_lineal_teori():
    return render_template('busqueda_lineal_teori.html') 

@app.route('/linear_search', methods=['POST'])
def linear_search():
    data = request.json
    array = data['array']
    target = data['target']

    start_time = time.time()
    indexes = [i for i, x in enumerate(array) if x == target]
    end_time = time.time()
    time_elapsed = end_time - start_time

    return jsonify({
        'array': array,
        'indexes': indexes,
        'time_elapsed': time_elapsed
    })

@app.route('/busqueda_binaria')
def busqueda_binaria():
    return render_template('busqueda_binaria.html')

@app.route('/busqueda_binaria_teori')
def busqueda_binaria_teori():
    return render_template('busqueda_binaria_teori.html') 

@app.route('/binary_search', methods=['POST'])
def binary_search():
    data = request.json
    array = sorted(data['array'])  # Ensure the array is sorted
    target = data['target']

    start_time = time.time()
    indexes = binary_search_indexes(array, target)
    end_time = time.time()
    time_elapsed = end_time - start_time

    return jsonify({
        'array': array,
        'indexes': indexes,
        'time_elapsed': time_elapsed
    })

def binary_search_indexes(arr, target):
    low, high = 0, len(arr) - 1
    result = []

    while low <= high:
        mid = (low + high) // 2
        if arr[mid] == target:
            result.append(mid)
            # Check for duplicates on the left side
            left = mid - 1
            while left >= 0 and arr[left] == target:
                result.append(left)
                left -= 1
            # Check for duplicates on the right side
            right = mid + 1
            while right < len(arr) and arr[right] == target:
                result.append(right)
                right += 1
            break
        elif arr[mid] < target:
            low = mid + 1
        else:
            high = mid - 1

    return sorted(result)



@app.route('/archivos_merge_sort/')
def list_archivos_merge_sort():
    files = os.listdir(app.config['carpeta_merge'])
    return render_template('lista_archivos_merge.html', files=files)

@app.route('/archivos_merge_sort/<filename>')
def download_file_merge(filename):
    # Descargar el archivo seleccionado
    return send_from_directory(app.config['carpeta_merge'], filename)

@app.route('/merge_archi')
def merge_archi():
    return render_template('merge_archi.html')
  
@app.route('/upload_merge_archiv', methods=['POST'])
def upload_merge_archiv():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['carpeta_archivos_subidos'], filename)
        file.save(filepath)

        filename, file_extension = os.path.splitext(file.filename)
        numbers = []

        if file_extension in ['.txt', '.csv']:
            with open(filepath, 'r') as f:
                data = f.read()
            numbers = list(map(int, data.split(',')))

        elif file_extension == '.xlsx':
            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, int):
                        numbers.append(cell)
        else:
            return jsonify({'error': 'Unsupported file type'}), 400

        sorting_thread = threading.Thread(target=merge_sort, args=(numbers,))
        sorting_thread.start()
        sorting_thread.join()

        sorted_filename = generate_unique_filename(app.config['carpeta_merge'], f"merge_sort_{filename}", file_extension[1:])
        sorted_filepath = os.path.join(app.config['carpeta_merge'], sorted_filename)

        if file_extension in ['.txt', '.csv']:
            with open(sorted_filepath, 'w') as f:
                f.write(','.join(map(str, numbers)))

        elif file_extension == '.xlsx':
            wb = Workbook()
            ws = wb.active
            for idx, num in enumerate(numbers, start=1):
                ws.cell(row=idx, column=1, value=num)
            wb.save(sorted_filepath)

        return jsonify({
            'success': True,
            'download_url': url_for('download_file', filename=sorted_filename)
        })

@app.route('/merge')
def merge():
    return render_template('merge.html')

@app.route('/merge_teori')
def merge_teori():
    return render_template('merge_teori.html')

@app.route('/merge_gen_ale')
def merge_gen_ale():
    return render_template('merge_gen_ale.html')

fake = Faker()

@app.route('/upload_merge_gen_ale', methods=['POST'])
def upload_merge_gen_ale():
    tipo_datos = request.form['tipo_datos']
    cantidad = int(request.form['cantidad_date'])

    if tipo_datos == 'numeros_enteros':
        generated_data = [random.randint(0, 100) for _ in range(cantidad)]
    elif tipo_datos == 'numeros_decimales':
        generated_data = [round(random.uniform(0, 100), 2) for _ in range(cantidad)]
    elif tipo_datos == 'nombres':
        generated_data = [fake.first_name() for _ in range(cantidad)]
    elif tipo_datos == 'palabras':
        generated_data = [fake.word() for _ in range(cantidad)]
    else:
        return jsonify({'error': 'Unsupported data type'}), 400

    return jsonify({'success': True, 'generated_data': generated_data})

@app.route('/ordenacion_merge_datos_aleatorios', methods=['POST'])
def ordenacion_merge_datos_aleatorios():
    data = request.json['data']
    # Ordenar los datos usando el algoritmo de ordenación burbuja
    sorted_data = merge_sort(data)

    # Generar un nombre de archivo único
    base_filename = 'datos_aletorios_ordenados_merge'
    sorted_filename = generate_unique_filename(app.config['carpeta_merge'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_merge'], sorted_filename)

    # Guardar los datos ordenados en un archivo
    with open(sorted_filepath, 'w') as f:
        f.write(','.join(map(str, sorted_data)))

    # Devolver la respuesta con los datos ordenados y la URL de descarga
    return jsonify({'success': True, 'sorted_data': sorted_data, 'download_url': url_for('download_file', filename=sorted_filename)})

@app.route('/merge_manual')
def merge_manual():
    return render_template('merge_manual.html')   

@app.route('/merge_manual_form', methods=['POST'])
def merge_manual_form():
    tipo_datos = request.form['tipo_datos']
    datos_a_ordenar = request.form['date_manu']
    separador = request.form['separador']

    if separador == 'comas':
        datos_list = datos_a_ordenar.split(',')
    elif separador == 'espacios':
        datos_list = datos_a_ordenar.split()

    if tipo_datos == 'numeros_enteros':
        datos_list = list(map(int, datos_list))
    elif tipo_datos == 'palabras':
        datos_list = list(map(str, datos_list))

    sorting_thread = threading.Thread(target=merge_sort, args=(datos_list,))
    sorting_thread.start()
    sorting_thread.join()

    sorted_list = datos_list
    sorted_data = ', '.join(map(str, sorted_list)) if separador == 'comas' else ' '.join(map(str, sorted_list))

    # Generar un nombre de archivo único
    base_filename = 'datos_manualmente_ordenados'
    sorted_filename = generate_unique_filename(app.config['carpeta_merge'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_merge'], sorted_filename)

    with open(sorted_filepath, 'w') as f:
        f.write(sorted_data)

    return jsonify({
        'sorted_data': sorted_data,
        'download_url': url_for('download_file', filename=sorted_filename)
    })

#--------------- HEAP SORT ------- HEAP SORT -----------

@app.route('/archivos_heap_sort/')
def list_archivos_heap_sort():
    files = os.listdir(app.config['carpeta_heap'])
    return render_template('lista_archivos_heap.html', files=files)

@app.route('/archivos_heap_sort/<filename>')
def download_file_heap(filename):
    # Descargar el archivo seleccionado
    return send_from_directory(app.config['carpeta_heap'], filename)

@app.route('/heap_archi')
def heap_archi():
    return render_template('heap_archi.html')
  
@app.route('/upload_heap_archiv', methods=['POST'])
def upload_heap_archiv():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['carpeta_archivos_subidos'], filename)
        file.save(filepath)

        filename, file_extension = os.path.splitext(file.filename)
        numbers = []

        if file_extension in ['.txt', '.csv']:
            with open(filepath, 'r') as f:
                data = f.read()
            numbers = list(map(int, data.split(',')))

        elif file_extension == '.xlsx':
            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, int):
                        numbers.append(cell)
        else:
            return jsonify({'error': 'Unsupported file type'}), 400

        sorting_thread = threading.Thread(target=heap_sort, args=(numbers,))
        sorting_thread.start()
        sorting_thread.join()

        sorted_filename = generate_unique_filename(app.config['carpeta_heap'], f"heap_sort_{filename}", file_extension[1:])
        sorted_filepath = os.path.join(app.config['carpeta_heap'], sorted_filename)

        if file_extension in ['.txt', '.csv']:
            with open(sorted_filepath, 'w') as f:
                f.write(','.join(map(str, numbers)))

        elif file_extension == '.xlsx':
            wb = Workbook()
            ws = wb.active
            for idx, num in enumerate(numbers, start=1):
                ws.cell(row=idx, column=1, value=num)
            wb.save(sorted_filepath)

        return jsonify({
            'success': True,
            'download_url': url_for('download_file', filename=sorted_filename)
        })

@app.route('/heap')
def heap():
    return render_template('heap.html')

@app.route('/heap_teori')
def heap_teori():
    return render_template('heap_teori.html')

@app.route('/heap_gen_ale')
def heap_gen_ale():
    return render_template('heap_gen_ale.html')

fake = Faker()

@app.route('/upload_heap_gen_ale', methods=['POST'])
def upload_heap_gen_ale():
    tipo_datos = request.form['tipo_datos']
    cantidad = int(request.form['cantidad_date'])

    if tipo_datos == 'numeros_enteros':
        generated_data = [random.randint(0, 100) for _ in range(cantidad)]
    elif tipo_datos == 'numeros_decimales':
        generated_data = [round(random.uniform(0, 100), 2) for _ in range(cantidad)]
    elif tipo_datos == 'nombres':
        generated_data = [fake.first_name() for _ in range(cantidad)]
    elif tipo_datos == 'palabras':
        generated_data = [fake.word() for _ in range(cantidad)]
    else:
        return jsonify({'error': 'Unsupported data type'}), 400

    return jsonify({'success': True, 'generated_data': generated_data})

@app.route('/ordenacion_heap_datos_aleatorios', methods=['POST'])
def ordenacion_heap_datos_aleatorios():
    data = request.json['data']
    # Ordenar los datos usando el algoritmo de ordenación burbuja
    sorted_data = heap_sort(data)

    # Generar un nombre de archivo único
    base_filename = 'datos_aletorios_ordenados_heap'
    sorted_filename = generate_unique_filename(app.config['carpeta_heap'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_heap'], sorted_filename)

    # Guardar los datos ordenados en un archivo
    with open(sorted_filepath, 'w') as f:
        f.write(','.join(map(str, sorted_data)))

    # Devolver la respuesta con los datos ordenados y la URL de descarga
    return jsonify({'success': True, 'sorted_data': sorted_data, 'download_url': url_for('download_file', filename=sorted_filename)})

@app.route('/heap_manual')
def heap_manual():
    return render_template('heap_manual.html')   

@app.route('/heap_manual_form', methods=['POST'])
def heap_manual_form():
    tipo_datos = request.form['tipo_datos']
    datos_a_ordenar = request.form['date_manu']
    separador = request.form['separador']

    if separador == 'comas':
        datos_list = datos_a_ordenar.split(',')
    elif separador == 'espacios':
        datos_list = datos_a_ordenar.split()

    if tipo_datos == 'numeros_enteros':
        datos_list = list(map(int, datos_list))
    elif tipo_datos == 'palabras':
        datos_list = list(map(str, datos_list))

    sorting_thread = threading.Thread(target=heap_sort, args=(datos_list,))
    sorting_thread.start()
    sorting_thread.join()

    sorted_list = datos_list
    sorted_data = ', '.join(map(str, sorted_list)) if separador == 'comas' else ' '.join(map(str, sorted_list))

    # Generar un nombre de archivo único
    base_filename = 'datos_manualmente_ordenados'
    sorted_filename = generate_unique_filename(app.config['carpeta_heap'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_heap'], sorted_filename)

    with open(sorted_filepath, 'w') as f:
        f.write(sorted_data)

    return jsonify({
        'sorted_data': sorted_data,
        'download_url': url_for('download_file', filename=sorted_filename)
    })

 # COUNTING SORT ------------------------------- COUNTING SORT

@app.route('/archivos_counting_sort/')
def list_archivos_counting_sort():
    files = os.listdir(app.config['carpeta_counting'])
    return render_template('lista_archivos_counting.html', files=files)

@app.route('/archivos_counting_sort/<filename>')
def download_file_counting(filename):
    # Descargar el archivo seleccionado
    return send_from_directory(app.config['carpeta_counting'], filename)

@app.route('/counting_archi')
def counting_archi():
    return render_template('counting_archi.html')
  
@app.route('/upload_counting_archiv', methods=['POST'])
def upload_counting_archiv():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['carpeta_archivos_subidos'], filename)
        file.save(filepath)

        filename, file_extension = os.path.splitext(file.filename)
        numbers = []

        if file_extension in ['.txt', '.csv']:
            with open(filepath, 'r') as f:
                data = f.read()
            numbers = list(map(int, data.split(',')))

        elif file_extension == '.xlsx':
            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, int):
                        numbers.append(cell)
        else:
            return jsonify({'error': 'Unsupported file type'}), 400

        sorting_thread = threading.Thread(target=counting_sort, args=(numbers,))
        sorting_thread.start()
        sorting_thread.join()

        sorted_filename = generate_unique_filename(app.config['carpeta_counting'], f"counting_sort_{filename}", file_extension[1:])
        sorted_filepath = os.path.join(app.config['carpeta_counting'], sorted_filename)

        if file_extension in ['.txt', '.csv']:
            with open(sorted_filepath, 'w') as f:
                f.write(','.join(map(str, numbers)))

        elif file_extension == '.xlsx':
            wb = Workbook()
            ws = wb.active
            for idx, num in enumerate(numbers, start=1):
                ws.cell(row=idx, column=1, value=num)
            wb.save(sorted_filepath)

        return jsonify({
            'success': True,
            'download_url': url_for('download_file', filename=sorted_filename)
        })

@app.route('/counting')
def counting():
    return render_template('counting.html')

@app.route('/counting_teori')
def counting_teori():
    return render_template('counting_teori.html')

@app.route('/counting_gen_ale')
def counting_gen_ale():
    return render_template('counting_gen_ale.html')

fake = Faker()

@app.route('/upload_counting_gen_ale', methods=['POST'])
def upload_counting_gen_ale():
    tipo_datos = request.form['tipo_datos']
    cantidad = int(request.form['cantidad_date'])

    if tipo_datos == 'numeros_enteros':
        generated_data = [random.randint(0, 100) for _ in range(cantidad)]
    elif tipo_datos == 'numeros_decimales':
        generated_data = [round(random.uniform(0, 100), 2) for _ in range(cantidad)]
    elif tipo_datos == 'nombres':
        generated_data = [fake.first_name() for _ in range(cantidad)]
    elif tipo_datos == 'palabras':
        generated_data = [fake.word() for _ in range(cantidad)]
    else:
        return jsonify({'error': 'Unsupported data type'}), 400

    return jsonify({'success': True, 'generated_data': generated_data})

@app.route('/ordenacion_counting_datos_aleatorios', methods=['POST'])
def ordenacion_counting_datos_aleatorios():
    data = request.json['data']
    # Ordenar los datos usando el algoritmo de ordenación burbuja
    sorted_data = counting_sort(data)

    # Generar un nombre de archivo único
    base_filename = 'datos_aletorios_ordenados_counting'
    sorted_filename = generate_unique_filename(app.config['carpeta_counting'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_counting'], sorted_filename)

    # Guardar los datos ordenados en un archivo
    with open(sorted_filepath, 'w') as f:
        f.write(','.join(map(str, sorted_data)))

    # Devolver la respuesta con los datos ordenados y la URL de descarga
    return jsonify({'success': True, 'sorted_data': sorted_data, 'download_url': url_for('download_file', filename=sorted_filename)})

@app.route('/counting_manual')
def counting_manual():
    return render_template('counting_manual.html')   

@app.route('/counting_manual_form', methods=['POST'])
def counting_manual_form():
    tipo_datos = request.form['tipo_datos']
    datos_a_ordenar = request.form['date_manu']
    separador = request.form['separador']

    if separador == 'comas':
        datos_list = datos_a_ordenar.split(',')
    elif separador == 'espacios':
        datos_list = datos_a_ordenar.split()

    if tipo_datos == 'numeros_enteros':
        datos_list = list(map(int, datos_list))
    elif tipo_datos == 'palabras':
        datos_list = list(map(str, datos_list))

    sorting_thread = threading.Thread(target=counting_sort, args=(datos_list,))
    sorting_thread.start()
    sorting_thread.join()

    sorted_list = datos_list
    sorted_data = ', '.join(map(str, sorted_list)) if separador == 'comas' else ' '.join(map(str, sorted_list))

    # Generar un nombre de archivo único
    base_filename = 'datos_manualmente_ordenados'
    sorted_filename = generate_unique_filename(app.config['carpeta_counting'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_counting'], sorted_filename)

    with open(sorted_filepath, 'w') as f:
        f.write(sorted_data)

    return jsonify({
        'sorted_data': sorted_data,
        'download_url': url_for('download_file', filename=sorted_filename)
    })

# ------------- RADIX SORT ------------ RADIX SORT-----------
@app.route('/archivos_radix_sort/')
def list_archivos_radix_sort():
    files = os.listdir(app.config['carpeta_radix'])
    return render_template('lista_archivos_radix.html', files=files)

@app.route('/archivos_radix_sort/<filename>')
def download_file_radix(filename):
    return send_from_directory(app.config['carpeta_radix'], filename)

@app.route('/radix_archi')
def radix_archi():
    return render_template('radix_archi.html')
  
@app.route('/upload_radix_archiv', methods=['POST'])
def upload_radix_archiv():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['carpeta_archivos_subidos'], filename)
        file.save(filepath)

        filename, file_extension = os.path.splitext(file.filename)
        numbers = []

        if file_extension in ['.txt', '.csv']:
            with open(filepath, 'r') as f:
                data = f.read()
            numbers = list(map(int, data.split(',')))

        elif file_extension == '.xlsx':
            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, int):
                        numbers.append(cell)
        else:
            return jsonify({'error': 'Unsupported file type'}), 400

        sorting_thread = threading.Thread(target=counting_sort, args=(numbers,))
        sorting_thread.start()
        sorting_thread.join()

        sorted_filename = generate_unique_filename(app.config['carpeta_radix'], f"radix_sort_{filename}", file_extension[1:])
        sorted_filepath = os.path.join(app.config['carpeta_radix'], sorted_filename)

        if file_extension in ['.txt', '.csv']:
            with open(sorted_filepath, 'w') as f:
                f.write(','.join(map(str, numbers)))

        elif file_extension == '.xlsx':
            wb = Workbook()
            ws = wb.active
            for idx, num in enumerate(numbers, start=1):
                ws.cell(row=idx, column=1, value=num)
            wb.save(sorted_filepath)

        return jsonify({
            'success': True,
            'download_url': url_for('download_file', filename=sorted_filename)
        })
    
@app.route('/upload_radix_gen_ale', methods=['POST'])
def upload_radix_gen_ale():
    tipo_datos = request.form['tipo_datos']
    cantidad = int(request.form['cantidad_date'])

    if tipo_datos == 'numeros_enteros':
        generated_data = [random.randint(0, 100) for _ in range(cantidad)]
    elif tipo_datos == 'numeros_decimales':
        generated_data = [round(random.uniform(0, 100), 2) for _ in range(cantidad)]
    elif tipo_datos == 'nombres':
        generated_data = [fake.first_name() for _ in range(cantidad)]
    elif tipo_datos == 'palabras':
        generated_data = [fake.word() for _ in range(cantidad)]
    else:
        return jsonify({'error': 'Unsupported data type'}), 400

    return jsonify({'success': True, 'generated_data': generated_data})

@app.route('/ordenacion_radix_datos_aleatorios', methods=['POST'])
def ordenacion_radix_datos_aleatorios():
    data = request.json['data']
    # Ordenar los datos usando el algoritmo de ordenación burbuja
    sorted_data = counting_sort(data)

    # Generar un nombre de archivo único
    base_filename = 'datos_aletorios_ordenados_counting'
    sorted_filename = generate_unique_filename(app.config['carpeta_radix'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_radix'], sorted_filename)

    # Guardar los datos ordenados en un archivo
    with open(sorted_filepath, 'w') as f:
        f.write(','.join(map(str, sorted_data)))

    # Devolver la respuesta con los datos ordenados y la URL de descarga
    return jsonify({'success': True, 'sorted_data': sorted_data, 'download_url': url_for('download_file', filename=sorted_filename)})

@app.route('/radix')
def radix():
    return render_template('radix.html')

@app.route('/radix_teori')
def radix_teori():
    return render_template('radix_teori.html')

@app.route('/radix_gen_ale')
def radix_gen_ale():
    return render_template('radix_gen_ale.html')

@app.route('/radix_manual')
def radix_manual():
    return render_template('radix_manual.html')   

@app.route('/radix_manual_form', methods=['POST'])
def radix_manual_form():
    tipo_datos = request.form['tipo_datos']
    datos_a_ordenar = request.form['date_manu']
    separador = request.form['separador']

    if separador == 'comas':
        datos_list = datos_a_ordenar.split(',')
    elif separador == 'espacios':
        datos_list = datos_a_ordenar.split()

    if tipo_datos == 'numeros_enteros':
        datos_list = list(map(int, datos_list))
    elif tipo_datos == 'palabras':
        datos_list = list(map(str, datos_list))

    sorting_thread = threading.Thread(target=radix_sort, args=(datos_list,))
    sorting_thread.start()
    sorting_thread.join()

    sorted_list = datos_list
    sorted_data = ', '.join(map(str, sorted_list)) if separador == 'comas' else ' '.join(map(str, sorted_list))

    # Generar un nombre de archivo único
    base_filename = 'datos_manualmente_ordenados'
    sorted_filename = generate_unique_filename(app.config['carpeta_radix'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_radix'], sorted_filename)

    with open(sorted_filepath, 'w') as f:
        f.write(sorted_data)

    return jsonify({
        'sorted_data': sorted_data,
        'download_url': url_for('download_file', filename=sorted_filename)
    })

# -------- QUICK SORT ------- QUICK SORT------------

@app.route('/archivos_quick_sort/')
def list_archivos_quick_sort():
    files = os.listdir(app.config['carpeta_quick'])
    return render_template('lista_archivos_quick.html', files=files)

@app.route('/archivos_quick_sort/<filename>')
def download_file_quick(filename):
    return send_from_directory(app.config['carpeta_quick'], filename)

@app.route('/quick_archi')
def quick_archi():
    return render_template('quick_archi.html')
  
@app.route('/upload_quick_archiv', methods=['POST'])
def upload_quick_archiv():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['carpeta_archivos_subidos'], filename)
        file.save(filepath)

        filename, file_extension = os.path.splitext(file.filename)
        numbers = []

        if file_extension in ['.txt', '.csv']:
            with open(filepath, 'r') as f:
                data = f.read()
            numbers = list(map(int, data.split(',')))

        elif file_extension == '.xlsx':
            wb = load_workbook(filepath)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, int):
                        numbers.append(cell)
        else:
            return jsonify({'error': 'Unsupported file type'}), 400

        sorting_thread = threading.Thread(target=quick_sort, args=(numbers,))
        sorting_thread.start()
        sorting_thread.join()

        sorted_filename = generate_unique_filename(app.config['carpeta_quick'], f"quick_sort_{filename}", file_extension[1:])
        sorted_filepath = os.path.join(app.config['carpeta_quick'], sorted_filename)

        if file_extension in ['.txt', '.csv']:
            with open(sorted_filepath, 'w') as f:
                f.write(','.join(map(str, numbers)))

        elif file_extension == '.xlsx':
            wb = Workbook()
            ws = wb.active
            for idx, num in enumerate(numbers, start=1):
                ws.cell(row=idx, column=1, value=num)
            wb.save(sorted_filepath)

        return jsonify({
            'success': True,
            'download_url': url_for('download_file', filename=sorted_filename)
        })
    
@app.route('/upload_quick_gen_ale', methods=['POST'])
def upload_quick_gen_ale():
    tipo_datos = request.form['tipo_datos']
    cantidad = int(request.form['cantidad_date'])

    if tipo_datos == 'numeros_enteros':
        generated_data = [random.randint(0, 100) for _ in range(cantidad)]
    elif tipo_datos == 'numeros_decimales':
        generated_data = [round(random.uniform(0, 100), 2) for _ in range(cantidad)]
    elif tipo_datos == 'nombres':
        generated_data = [fake.first_name() for _ in range(cantidad)]
    elif tipo_datos == 'palabras':
        generated_data = [fake.word() for _ in range(cantidad)]
    else:
        return jsonify({'error': 'Unsupported data type'}), 400

    return jsonify({'success': True, 'generated_data': generated_data})

@app.route('/ordenacion_quick_datos_aleatorios', methods=['POST'])
def ordenacion_quick_datos_aleatorios():
    data = request.json['data']
    # Ordenar los datos usando el algoritmo de ordenación burbuja
    sorted_data = quick_sort(data)

    # Generar un nombre de archivo único
    base_filename = 'datos_aletorios_ordenados_quick'
    sorted_filename = generate_unique_filename(app.config['carpeta_quick'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_quick'], sorted_filename)

    # Guardar los datos ordenados en un archivo
    with open(sorted_filepath, 'w') as f:
        f.write(','.join(map(str, sorted_data)))

    # Devolver la respuesta con los datos ordenados y la URL de descarga
    return jsonify({'success': True, 'sorted_data': sorted_data, 'download_url': url_for('download_file', filename=sorted_filename)})

@app.route('/quick')
def quick():
    return render_template('quick.html')

@app.route('/quick_teori')
def quick_teori():
    return render_template('quick_teori.html')

@app.route('/quick_gen_ale')
def quick_gen_ale():
    return render_template('quick_gen_ale.html')

@app.route('/quick_manual')
def quick_manual():
    return render_template('quick_manual.html')   

@app.route('/quick_manual_form', methods=['POST'])
def quick_manual_form():
    tipo_datos = request.form['tipo_datos']
    datos_a_ordenar = request.form['date_manu']
    separador = request.form['separador']

    if separador == 'comas':
        datos_list = datos_a_ordenar.split(',')
    elif separador == 'espacios':
        datos_list = datos_a_ordenar.split()

    if tipo_datos == 'numeros_enteros':
        datos_list = list(map(int, datos_list))
    elif tipo_datos == 'palabras':
        datos_list = list(map(str, datos_list))

    sorting_thread = threading.Thread(target=quick_sort, args=(datos_list,))
    sorting_thread.start()
    sorting_thread.join()

    sorted_list = datos_list
    sorted_data = ', '.join(map(str, sorted_list)) if separador == 'comas' else ' '.join(map(str, sorted_list))

    # Generar un nombre de archivo único
    base_filename = 'datos_manualmente_ordenados'
    sorted_filename = generate_unique_filename(app.config['carpeta_quick'], base_filename, 'txt')
    sorted_filepath = os.path.join(app.config['carpeta_quick'], sorted_filename)

    with open(sorted_filepath, 'w') as f:
        f.write(sorted_data)

    return jsonify({
        'sorted_data': sorted_data,
        'download_url': url_for('download_file', filename=sorted_filename)
    })

#hasta aqui hacen su parte 



@app.route('/about')
def about():
    return render_template('sobre_nosotros.html')

if __name__ == '__main__':
    app.run(debug=True)
